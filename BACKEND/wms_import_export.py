"""
SAMILA WMS - Import/Export Service
Handle Excel import and PDF export
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
import csv
import io
from datetime import datetime
from typing import List, Dict, Optional

# =============== EXCEL IMPORT ===============

class ExcelImportService:
    """Import data from Excel files"""
    
    @staticmethod
    def import_products(file_path: str) -> Dict:
        """Import products from Excel"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            
            products = []
            errors = []
            
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
                try:
                    if not row[0]:  # Skip empty rows
                        continue
                    
                    product = {
                        "sku": str(row[0]).strip(),
                        "barcode": str(row[1]).strip() if row[1] else None,
                        "name_en": str(row[2]).strip(),
                        "name_th": str(row[3]).strip(),
                        "category": str(row[4]).strip(),
                        "unit": str(row[5]).strip(),
                        "weight_kg": float(row[6]) if row[6] else 0,
                        "volume_m3": float(row[7]) if row[7] else 0,
                        "price": float(row[8]) if row[8] else 0
                    }
                    
                    # Validate required fields
                    if not product["sku"] or not product["name_en"]:
                        errors.append(f"Row {idx}: Missing SKU or Name")
                        continue
                    
                    # Check for duplicate SKU
                    if any(p["sku"] == product["sku"] for p in products):
                        errors.append(f"Row {idx}: Duplicate SKU {product['sku']}")
                        continue
                    
                    products.append(product)
                
                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")
            
            return {
                "status": "success",
                "imported_count": len(products),
                "error_count": len(errors),
                "errors": errors,
                "data": products
            }
        
        except Exception as e:
            return {
                "status": "error",
                "message": str(e)
            }
    
    @staticmethod
    def import_receiving_orders(file_path: str) -> Dict:
        """Import receiving orders from Excel"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            
            orders = []
            items = []
            errors = []
            
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
                try:
                    if not row[0]:
                        continue
                    
                    order = {
                        "po_number": str(row[0]).strip(),
                        "supplier_code": str(row[1]).strip(),
                        "expected_date": str(row[2]).strip(),
                        "sku": str(row[3]).strip(),
                        "quantity": float(row[4]),
                        "location": str(row[5]).strip()
                    }
                    
                    orders.append(order)
                
                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")
            
            return {
                "status": "success",
                "imported_count": len(orders),
                "error_count": len(errors),
                "errors": errors,
                "data": orders
            }
        
        except Exception as e:
            return {
                "status": "error",
                "message": str(e)
            }
    
    @staticmethod
    def import_sales_orders(file_path: str) -> Dict:
        """Import sales orders from Excel"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            
            orders = []
            errors = []
            
            for idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
                try:
                    if not row[0]:
                        continue
                    
                    order = {
                        "customer_code": str(row[0]).strip(),
                        "so_number": str(row[1]).strip(),
                        "order_date": str(row[2]).strip(),
                        "sku": str(row[3]).strip(),
                        "quantity": float(row[4]),
                        "delivery_address": str(row[5]).strip()
                    }
                    
                    orders.append(order)
                
                except Exception as e:
                    errors.append(f"Row {idx}: {str(e)}")
            
            return {
                "status": "success",
                "imported_count": len(orders),
                "error_count": len(errors),
                "errors": errors,
                "data": orders
            }
        
        except Exception as e:
            return {
                "status": "error",
                "message": str(e)
            }


# =============== EXCEL EXPORT ===============

class ExcelExportService:
    """Export data to Excel files"""
    
    @staticmethod
    def export_inventory(warehouse_id: int, inventory_data: List[Dict]) -> bytes:
        """Export inventory to Excel"""
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Inventory"
            
            # Headers
            headers = ["SKU", "Product", "Warehouse", "Location", "Quantity", 
                      "Available", "Reserved", "Status", "Last Updated"]
            worksheet.append(headers)
            
            # Style headers
            header_fill = PatternFill(start_color="00A8CC", end_color="00A8CC", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Data
            for item in inventory_data:
                worksheet.append([
                    item.get("sku"),
                    item.get("product"),
                    item.get("warehouse"),
                    item.get("location"),
                    item.get("quantity_on_hand"),
                    item.get("quantity_available"),
                    item.get("quantity_reserved"),
                    item.get("status"),
                    item.get("last_updated")
                ])
            
            # Adjust column widths
            worksheet.column_dimensions["A"].width = 12
            worksheet.column_dimensions["B"].width = 25
            worksheet.column_dimensions["C"].width = 15
            worksheet.column_dimensions["D"].width = 12
            worksheet.column_dimensions["E"].width = 12
            worksheet.column_dimensions["F"].width = 12
            worksheet.column_dimensions["G"].width = 12
            worksheet.column_dimensions["H"].width = 10
            
            # Save to bytes
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return output.getvalue()
        
        except Exception as e:
            raise Exception(f"Excel export error: {str(e)}")
    
    @staticmethod
    def export_receiving_orders(gr_data: List[Dict]) -> bytes:
        """Export receiving orders to Excel"""
        try:
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Receiving"
            
            headers = ["GR Number", "PO Number", "Supplier", "Date", "Receiver", 
                      "Items", "Pallets", "Status"]
            worksheet.append(headers)
            
            for item in gr_data:
                worksheet.append([
                    item.get("gr_number"),
                    item.get("po_number"),
                    item.get("supplier"),
                    item.get("date"),
                    item.get("receiver"),
                    item.get("items"),
                    item.get("pallets"),
                    item.get("status")
                ])
            
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            
            return output.getvalue()
        
        except Exception as e:
            raise Exception(f"Excel export error: {str(e)}")


# =============== PDF EXPORT ===============

class PDFExportService:
    """Export data to PDF"""
    
    @staticmethod
    def export_picking_list(pick_data: Dict) -> bytes:
        """Export picking list to PDF"""
        try:
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=letter, topMargin=0.5*inch)
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('00A8CC'),
                spaceAfter=12
            )
            elements.append(Paragraph("PICKING LIST", title_style))
            
            # Header info
            info_data = [
                ["Pick Number:", pick_data.get("pick_number")],
                ["Date:", datetime.now().strftime("%Y-%m-%d %H:%M")],
                ["Status:", pick_data.get("status")]
            ]
            
            info_table = Table(info_data, colWidths=[2*inch, 3*inch])
            info_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('E8F5E9')),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ]))
            
            elements.append(info_table)
            elements.append(Spacer(1, 12))
            
            # Items table
            items_data = [["SKU", "Product", "Location", "Qty", "Picked", "Status"]]
            
            for item in pick_data.get("items", []):
                items_data.append([
                    item.get("sku"),
                    item.get("product"),
                    item.get("location"),
                    str(item.get("quantity")),
                    str(item.get("picked", 0)),
                    item.get("status", "PENDING")
                ])
            
            items_table = Table(items_data, colWidths=[1.2*inch, 2*inch, 1.2*inch, 0.8*inch, 0.8*inch, 1*inch])
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('00A8CC')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('F5F5F5')])
            ]))
            
            elements.append(items_table)
            
            # Build PDF
            doc.build(elements)
            output.seek(0)
            
            return output.getvalue()
        
        except Exception as e:
            raise Exception(f"PDF export error: {str(e)}")
    
    @staticmethod
    def export_shipping_manifest(shipment_data: Dict) -> bytes:
        """Export shipping manifest to PDF"""
        try:
            output = io.BytesIO()
            doc = SimpleDocTemplate(output, pagesize=landscape(letter))
            elements = []
            styles = getSampleStyleSheet()
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.HexColor('00A8CC')
            )
            elements.append(Paragraph("SHIPPING MANIFEST", title_style))
            elements.append(Spacer(1, 12))
            
            # Manifest info
            manifest_data = [
                ["Shipment Number:", shipment_data.get("shipment_number"), 
                 "Carrier:", shipment_data.get("carrier")],
                ["Date:", datetime.now().strftime("%Y-%m-%d"), 
                 "Tracking:", shipment_data.get("tracking_number")],
                ["Weight:", f"{shipment_data.get('weight')} kg", 
                 "Pallets:", shipment_data.get("pallets")]
            ]
            
            manifest_table = Table(manifest_data, colWidths=[2*inch, 2.5*inch, 2*inch, 2.5*inch])
            manifest_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
            ]))
            
            elements.append(manifest_table)
            elements.append(Spacer(1, 12))
            
            # Items table
            items_data = [["Box", "SKU", "Product", "Quantity", "Seal", "Signature"]]
            
            for item in shipment_data.get("items", []):
                items_data.append([
                    item.get("box_number"),
                    item.get("sku"),
                    item.get("product"),
                    str(item.get("quantity")),
                    "",
                    ""
                ])
            
            items_table = Table(items_data, colWidths=[1*inch, 1.2*inch, 2.5*inch, 1.2*inch, 1.5*inch, 1.5*inch])
            items_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('00A8CC')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('F5F5F5')])
            ]))
            
            elements.append(items_table)
            
            # Build PDF
            doc.build(elements)
            output.seek(0)
            
            return output.getvalue()
        
        except Exception as e:
            raise Exception(f"PDF export error: {str(e)}")


# =============== CSV EXPORT ===============

class CSVExportService:
    """Export data to CSV"""
    
    @staticmethod
    def export_inventory_csv(inventory_data: List[Dict]) -> str:
        """Export inventory to CSV"""
        try:
            output = io.StringIO()
            fieldnames = ["SKU", "Product", "Warehouse", "Location", "Quantity", 
                         "Available", "Reserved", "Status"]
            
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            
            for item in inventory_data:
                writer.writerow({
                    "SKU": item.get("sku"),
                    "Product": item.get("product"),
                    "Warehouse": item.get("warehouse"),
                    "Location": item.get("location"),
                    "Quantity": item.get("quantity_on_hand"),
                    "Available": item.get("quantity_available"),
                    "Reserved": item.get("quantity_reserved"),
                    "Status": item.get("status")
                })
            
            return output.getvalue()
        
        except Exception as e:
            raise Exception(f"CSV export error: {str(e)}")
